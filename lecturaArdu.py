import serial
import openpyxl
from openpyxl import load_workbook
import time

# Configura el puerto serie (asegúrate de usar el puerto correcto)
ser = serial.Serial('COM4', 9600)  # Puerto

# Intenta cargar un archivo Excel existente, o crea uno nuevo si no existe
try:
    wb = load_workbook("datos_temperatura.xlsx")
    ws = wb.active
    print("Archivo existente cargado.")
except FileNotFoundError:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Datos de Temperatura"
    ws.append(["Timestamp", "Temperatura Promedio (C)"])
    print("Archivo nuevo creado.")

try:
    while True:
        if ser.in_waiting > 0:
            line = ser.readline().decode('utf-8').strip()
            print(f"Datos recibidos: {line}")  # Línea para depurar la entrada
            if "Promedio de temperatura" in line:
                # Extrae el valor de temperatura del mensaje
                temp_value = line.split(': ')[1].split(' ')[0]
                # Agrega una nueva fila en el Excel con la marca de tiempo y la temperatura
                ws.append([time.strftime("%Y-%m-%d %H:%M:%S"), temp_value])
                print(f"Datos guardados: {temp_value} °C")
                # Guarda el archivo Excel
                wb.save("datos_temperatura.xlsx")
except KeyboardInterrupt:
    # Guarda y cierra el archivo Excel al interrumpir el programa
    wb.save("datos_temperatura.xlsx")
    print("Programa terminado y archivo guardado.")
finally:
    ser.close()
